import pytest
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import openpyxl
from src.multi_tf_excel import build_multi_tf_excel


class TestMultiTfExcel:

    @pytest.fixture
    def sample_symbols(self):
        return ['RELIANCE', 'TCS', 'INFY']

    @pytest.fixture
    def sample_tf_list(self):
        return ['1m', '15m', '1d']

    @pytest.fixture
    def sample_dates(self):
        return datetime(2023, 1, 1), datetime(2023, 12, 31)

    def test_build_multi_tf_excel_creates_file(self, sample_symbols, sample_tf_list, sample_dates, tmp_path):
        """Test that build_multi_tf_excel creates an Excel file with expected sheets."""
        start, end = sample_dates
        out_path = tmp_path / "test_multi_tf.xlsx"

        # Mock the functions to avoid API calls
        import src.multi_tf_excel as mte
        original_fetch = mte.fetch_ohlc_v3
        original_resample = mte.resample_ohlc
        original_compute = mte.compute_indicators_for_tf
        original_select = mte.select_best_strategy

        def mock_fetch(symbols, tf, start, end):
            # Return mock data
            dates = pd.date_range(start, end, freq='1D')[:10]     
            data = []
            for symbol in symbols:
                for date in dates:
                    data.append({
                        'Symbol': symbol,
                        'Date': date,
                        'Open': 100,
                        'High': 105,
                        'Low': 95,
                        'Close': 102,
                        'Volume': 1000
                    })
            return pd.DataFrame(data)

        def mock_resample(df, tf):
            return df  # Return as is for simplicity

        def mock_compute(df, tf):
            df = df.copy()
            df['EMA20'] = df['Close']
            df['EMA50'] = df['Close']
            df['EMA200'] = df['Close']
            df['RSI14'] = 50
            df['ATR14'] = 1
            df['BB_MA20'] = df['Close']
            df['BB_Upper'] = 105
            df['BB_Lower'] = 95
            df['BB_BandWidth'] = 0.1
            df['DonchianH20'] = 105
            df['RVOL20'] = 1
            df['AVWAP60'] = 100
            df['KC_Upper'] = 105
            df['KC_Lower'] = 95
            df['Trend_OK'] = 1
            df['Minervini_Trend'] = 1
            df['RS_vs_Index'] = 1
            df['RS_ROC20'] = 0
            df['IndexUpRegime'] = 1
            return df

        def mock_select(candidates, strategies, cfg, out_dir):
            return {'selected': 'AVWAP', 'results': {}}

        mte.fetch_ohlc_v3 = mock_fetch
        mte.resample_ohlc = mock_resample
        mte.compute_indicators_for_tf = mock_compute
        mte.select_best_strategy = mock_select

        try:
            build_multi_tf_excel(sample_symbols, sample_tf_list, start, end, str(out_path))

            # Check file exists
            assert out_path.exists()

            # Check sheets
            wb = openpyxl.load_workbook(out_path)
            expected_sheets = [f"NIFTY50_{tf}" for tf in sample_tf_list]
            actual_sheets = wb.sheetnames
            for sheet in expected_sheets:
                assert sheet in actual_sheets

            # Check each sheet has data
            for sheet_name in expected_sheets:
                ws = wb[sheet_name]
                assert ws.max_row > 1  # Has header + data
                assert ws.max_column >= 10  # Has expected columns

        finally:
            # Restore originals
            mte.fetch_ohlc_v3 = original_fetch
            mte.resample_ohlc = original_resample
            mte.compute_indicators_for_tf = original_compute
            mte.select_best_strategy = original_select

# Mock the functions to avoid API calls
        import src.multi_tf_excel as mte
        original_fetch = mte.fetch_ohlc_v3
        original_resample = mte.resample_ohlc
        original_compute = mte.compute_indicators_for_tf
        original_select = mte.select_best_strategy

        def mock_fetch(symbols, tf, start, end):
            dates = pd.date_range(start, end, freq='1D')[:5]      
            data = []
            for symbol in symbols[:1]:  # One symbol
                for date in dates:
                    data.append({
                        'Symbol': symbol,
                        'Date': date,
                        'Open': 100,
                        'High': 105,
                        'Low': 95,
                        'Close': 102,
                        'Volume': 1000
                    })
            return pd.DataFrame(data)

        def mock_resample(df, tf):
            return df

        def mock_compute(df, tf):
            df = df.copy()
            df['EMA20'] = df['Close']
            df['EMA50'] = df['Close']
            df['EMA200'] = df['Close']
            df['RSI14'] = 50
            df['ATR14'] = 1
            df['BB_MA20'] = df['Close']
            df['BB_Upper'] = 105
            df['BB_Lower'] = 95
            df['BB_BandWidth'] = 0.1
            df['DonchianH20'] = 105
            df['RVOL20'] = 1
            df['AVWAP60'] = 100
            df['KC_Upper'] = 105
            df['KC_Lower'] = 95
            df['Trend_OK'] = 1
            df['Minervini_Trend'] = 1
            df['RS_vs_Index'] = 1
            df['RS_ROC20'] = 0
            df['IndexUpRegime'] = 1
            return df

        def mock_select(candidates, strategies, cfg, out_dir):
            return {'selected': 'AVWAP', 'results': {}}

        mte.fetch_ohlc_v3 = mock_fetch
        mte.resample_ohlc = mock_resample
        mte.compute_indicators_for_tf = mock_compute
        mte.select_best_strategy = mock_select

        try:
            build_multi_tf_excel(sample_symbols[:1], sample_tf_list[:1], start, end, str(out_path))

            wb = openpyxl.load_workbook(out_path)
            ws = wb.active
            header = [cell.value for cell in ws[1]]

            expected_columns = [
                'Symbol', 'GTT_Buy_Price', 'Stoploss', 'Sell_Rate', 'Strategy', 'Notes', 'Explanation',
                'RSI14', 'RSI14_Status', 'GoldenBull_Flag', 'GoldenBear_Flag', 'GoldenBull_Date', 'GoldenBear_Date',
                'Generated_At_IST'
            ]

            for col in expected_columns:
                assert col in header

        finally:
            mte.fetch_ohlc_v3 = original_fetch
            mte.resample_ohlc = original_resample
            mte.compute_indicators_for_tf = original_compute
            mte.select_best_strategy = original_select