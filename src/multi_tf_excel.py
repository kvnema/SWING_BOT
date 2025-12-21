import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import List
import pytz
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

from .timeframe_fetcher import fetch_ohlc_v3, resample_ohlc, compute_indicators_for_tf, latest_window
from .signals import compute_signals
from .scoring import compute_composite_score
from .select_strategy import select_best_strategy
from .gtt_sizing import build_gtt_plan
from .data_validation import load_metadata, validate_recency, validate_window, validate_symbols, get_today_ist


def build_multi_tf_excel(symbols: List[str], tf_list: List[str], start: datetime, end: datetime, out_xlsx: str, tz: str = "Asia/Kolkata") -> None:
    """
    Build multi-timeframe Excel workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for tf in tf_list:
        print(f"Processing {tf}...")

        # Determine base TF for fetching
        if tf in ['1m', '15m', '1h', '4h']:
            base_tf = 'I1'  # Fetch 1min and resample
        elif tf == '1d':
            base_tf = '1d'
        elif tf in ['1w', '1mo']:
            base_tf = '1d'  # Fetch daily and resample

        # Fetch base data
        if base_tf in ['I1', 'I30', '1d']:
            df_base = fetch_ohlc_v3(symbols, base_tf, start, end)
        else:
            raise ValueError(f"Unsupported base_tf: {base_tf}")

        if df_base.empty:
            print(f"No data for {tf}, skipping.")
            continue

        # Resample if needed
        if tf != base_tf:
            # Group by symbol and resample each
            resampled_frames = []
            for symbol in df_base['Symbol'].unique():
                df_sym = df_base[df_base['Symbol'] == symbol]
                df_res = resample_ohlc(df_sym, tf)
                resampled_frames.append(df_res)
            df_tf = pd.concat(resampled_frames, ignore_index=True)
        else:
            df_tf = df_base

        # Compute indicators
        processed_frames = []
        for symbol in df_tf['Symbol'].unique():
            df_sym = df_tf[df_tf['Symbol'] == symbol]
            df_ind = compute_indicators_for_tf(df_sym, tf)
            processed_frames.append(df_ind)
        df_indicators = pd.concat(processed_frames, ignore_index=True)

        # Compute signals
        df_signals = compute_signals(df_indicators)

        # Select top symbols (simplified: all with scores)
        df_signals['CompositeScore'] = compute_composite_score(df_signals)
        latest = df_signals.sort_values('Date').groupby('Symbol').tail(1)
        top_symbols = latest.nlargest(25, 'CompositeScore')['Symbol'].tolist()

        # Build GTT plan for top symbols
        candidates = df_signals[df_signals['Symbol'].isin(top_symbols)].copy()
        strategies = {
            'SEPA': 'SEPA_Flag',
            'VCP': 'VCP_Flag',
            'Donchian': 'Donchian_Breakout',
            'MR': 'MR_Flag',
            'Squeeze': 'SqueezeBreakout_Flag',
            'AVWAP': 'AVWAP_Reclaim_Flag'
        }
        selected = select_best_strategy(candidates, strategies, {'risk': {}, 'backtest': {}}, None)
        selected_strategy = selected['selected']

        # Build plan
        plan = build_gtt_plan(candidates, selected_strategy, {'risk': {}}, {})

        # Rename columns to match spec
        plan = plan.rename(columns={
            'ENTRY_trigger_price': 'GTT_Buy_Price',
            'STOPLOSS_trigger_price': 'Stoploss',
            'TARGET_trigger_price': 'Sell_Rate'
        })

        # Add columns as per spec
        plan['Strategy'] = selected_strategy
        plan['Notes'] = f"{selected_strategy} signal"
        plan['Explanation'] = f"Auto-generated for {tf}"
        plan['Generated_At_IST'] = datetime.now(pytz.timezone(tz)).strftime('%Y-%m-%d %H:%M:%S')

        # Select columns
        columns = [
            'Symbol', 'GTT_Buy_Price', 'Stoploss', 'Sell_Rate', 'Strategy', 'Notes', 'Explanation',
            'RSI14', 'RSI14_Status', 'GoldenBull_Flag', 'GoldenBear_Flag', 'GoldenBull_Date', 'GoldenBear_Date',
            'Generated_At_IST'
        ]
        plan = plan[columns]

        # Create sheet
        sheet_name = f"NIFTY50_{tf}"
        ws = wb.create_sheet(sheet_name)

        # Write data
        for r, row in enumerate(dataframe_to_rows(plan, index=False, header=True), 1):
            for c, value in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=value)
                if r == 1:
                    cell.font = Font(bold=True)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Save workbook
    wb.save(out_xlsx)
    print(f"Multi-TF Excel saved to {out_xlsx}")