"""
SWING_BOT Final Excel
====================

Generate final delivery Excel with GTT plans and confidence scores.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

from .utils import load_config, logger, get_ist_now

class FinalExcelGenerator:
    """Generate final delivery Excel with GTT plans and confidence."""

    def __init__(self, config: Optional[Dict] = None):
        """Initialize final Excel generator with configuration."""
        self.config = config or load_config()
        self.excel_config = self.config.get('final_excel', {})

        logger.info("Initialized FinalExcelGenerator")

    def load_gtt_plan(self, gtt_plan_path: str) -> pd.DataFrame:
        """Load GTT plan data."""
        try:
            df = pd.read_csv(gtt_plan_path)
            logger.info(f"Loaded GTT plan with {len(df)} positions")
            return df
        except Exception as e:
            logger.error(f"Failed to load GTT plan: {str(e)}")
            return pd.DataFrame()

    def load_backtest_results(self, backtest_dir: str) -> Dict:
        """Load backtest results for context."""
        backtest_results = {}

        try:
            # Load selected strategy
            selected_file = f"{backtest_dir}/selected_strategy.json"
            if os.path.exists(selected_file):
                import json
                with open(selected_file, 'r') as f:
                    backtest_results = json.load(f)
                logger.info("Loaded backtest results for final Excel")
        except Exception as e:
            logger.warning(f"Could not load backtest results: {str(e)}")

        return backtest_results

    def create_summary_sheet(self, wb: Workbook, gtt_df: pd.DataFrame,
                           backtest_results: Dict) -> None:
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "SWING_BOT GTT Delivery Plan"
        ws['A1'].font = Font(size=16, bold=True)

        ws['A2'] = f"Generated: {get_ist_now().strftime('%Y-%m-%d %H:%M:%S IST')}"
        ws['A2'].font = Font(italic=True)

        # Strategy information
        ws['A4'] = "Strategy Information:"
        ws['A4'].font = Font(bold=True)

        row = 5
        if backtest_results:
            ws[f'A{row}'] = f"Selected Strategy: {backtest_results.get('selected_strategy', 'N/A')}"
            ws[f'A{row+1}'] = f"Selection Score: {backtest_results.get('selection_score', 0):.2f}"
            ws[f'A{row+2}'] = f"Market Regime: {backtest_results.get('market_regime', 'N/A')}"

            if 'metrics' in backtest_results:
                metrics = backtest_results['metrics']
                ws[f'A{row+4}'] = f"Sharpe Ratio: {metrics.get('Sharpe_Ratio', 0):.2f}"
                ws[f'A{row+5}'] = f"Total Return: {metrics.get('Total_Return', 0):.1f}%"
                ws[f'A{row+6}'] = f"Win Rate: {metrics.get('Win_Rate', 0):.1f}%"

        # Portfolio summary
        ws['E4'] = "Portfolio Summary:"
        ws['E4'].font = Font(bold=True)

        if not gtt_df.empty:
            # Calculate total value from Qty * ENTRY_trigger_price
            total_value = (gtt_df['Qty'] * gtt_df['ENTRY_trigger_price']).sum()
            total_risk = gtt_df['Risk_Amount'].sum() if 'Risk_Amount' in gtt_df.columns else 0
            avg_confidence = gtt_df['DecisionConfidence'].mean()

            ws['E5'] = f"Total Positions: {len(gtt_df)}"
            ws['E6'] = f"Total Value: ₹{total_value:,.0f}"
            ws['E7'] = f"Total Risk: ₹{total_risk:,.0f}"
            ws['E8'] = f"Average Confidence: {avg_confidence:.1f}/5"

        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

    def create_gtt_plan_sheet(self, wb: Workbook, gtt_df: pd.DataFrame) -> None:
        """Create GTT plan details sheet."""
        ws = wb.create_sheet("GTT-Delivery-Plan")

        # Title
        ws['A1'] = "GTT Delivery Orders"
        ws['A1'].font = Font(size=14, bold=True)

        if gtt_df.empty:
            ws['A3'] = "No GTT plan data available"
            return

        # Specific column ordering as per requirements
        headers = [
            'Symbol', 'Qty', 'ENTRY_trigger_price', 'TARGET_trigger_price', 'STOPLOSS_trigger_price',
            'DecisionConfidence', 'Confidence_Level', 'R', 'Explanation', 'GTT_Explanation',
            'RSI14_D', 'RSI_Above50_D', 'RSI_Overbought_D', 'MACD_Line_D', 'MACD_Signal_D', 'MACD_Hist_D', 'MACD_CrossUp_D', 'MACD_AboveZero_D',
            'RSI14_H4', 'RSI_Above50_H4', 'MACD_CrossUp_H4',
            'RSI_MACD_Confirmations_OK', 'RSI_MACD_Notes',
            'Audit_Flag', 'Issues', 'Fix_Suggestion', 'Pivot_Source', 'Entry_Logic', 'Stop_Logic', 
            'Target_Logic', 'Latest_Close', 'Latest_LTP', 'Canonical_Entry', 'Canonical_Stop', 'Canonical_Target'
        ]

        # Write headers
        for j, header in enumerate(headers, 1):
            cell = ws[f'{get_column_letter(j)}2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data - write in specific column order
        for i, (_, row) in enumerate(gtt_df.iterrows(), 3):
            ws[f'A{i}'] = row.get('Symbol', '')
            ws[f'B{i}'] = row.get('Qty', 0)
            ws[f'C{i}'] = row.get('ENTRY_trigger_price', 0)
            ws[f'D{i}'] = row.get('TARGET_trigger_price', 0)
            ws[f'E{i}'] = row.get('STOPLOSS_trigger_price', 0)
            ws[f'F{i}'] = row.get('DecisionConfidence', 0)
            ws[f'G{i}'] = row.get('Confidence_Level', '')
            ws[f'H{i}'] = row.get('R', 0)
            ws[f'I{i}'] = row.get('Explanation', '')
            ws[f'J{i}'] = row.get('GTT_Explanation', '')
            ws[f'K{i}'] = row.get('RSI14', 0)  # RSI14_D
            ws[f'L{i}'] = row.get('RSI_Above50', False)  # RSI_Above50_D
            ws[f'M{i}'] = row.get('RSI_Overbought', False)  # RSI_Overbought_D
            ws[f'N{i}'] = row.get('MACD_Line', 0)  # MACD_Line_D
            ws[f'O{i}'] = row.get('MACD_Signal', 0)  # MACD_Signal_D
            ws[f'P{i}'] = row.get('MACD_Hist', 0)  # MACD_Hist_D
            ws[f'Q{i}'] = row.get('MACD_CrossUp', False)  # MACD_CrossUp_D
            ws[f'R{i}'] = row.get('MACD_AboveZero', False)  # MACD_AboveZero_D
            ws[f'S{i}'] = row.get('RSI14_H4', 0)  # RSI14_H4
            ws[f'T{i}'] = row.get('RSI_Above50_H4', False)  # RSI_Above50_H4
            ws[f'U{i}'] = row.get('MACD_CrossUp_H4', False)  # MACD_CrossUp_H4
            ws[f'V{i}'] = row.get('RSI_MACD_Confirmations_OK', False)
            ws[f'W{i}'] = row.get('RSI_MACD_Notes', '')
            ws[f'X{i}'] = row.get('Audit_Flag', '')
            ws[f'Y{i}'] = row.get('Issues', '')
            ws[f'Z{i}'] = row.get('Fix_Suggestion', '')
            ws[f'AA{i}'] = row.get('Pivot_Source', '')
            ws[f'AB{i}'] = row.get('Entry_Logic', '')
            ws[f'AC{i}'] = row.get('Stop_Logic', '')
            ws[f'AD{i}'] = row.get('Target_Logic', '')
            ws[f'AE{i}'] = row.get('Latest_Close', '')
            ws[f'AF{i}'] = row.get('Latest_LTP', '')
            ws[f'AG{i}'] = row.get('Canonical_Entry', '')
            ws[f'AH{i}'] = row.get('Canonical_Stop', '')
            ws[f'AI{i}'] = row.get('Canonical_Target', '')

        # Conditional formatting for confidence (column F)
        for i in range(3, len(gtt_df) + 3):
            confidence_cell = ws[f'F{i}']
            if confidence_cell.value >= 4:
                confidence_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green
            elif confidence_cell.value >= 3:
                confidence_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Yellow
            else:
                confidence_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Red
            
            # Conditional formatting for audit flag (column K)
            audit_cell = ws[f'K{i}']
            if audit_cell.value == 'FAIL':
                audit_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Red
            elif audit_cell.value == 'PASS':
                audit_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Green

        # Auto-adjust column widths
        for j in range(1, len(headers) + 1):
            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = 15

    def create_confidence_analysis_sheet(self, wb: Workbook, gtt_df: pd.DataFrame) -> None:
        """Create confidence analysis sheet."""
        ws = wb.create_sheet("Confidence_Analysis")

        # Title
        ws['A1'] = "Confidence Analysis"
        ws['A1'].font = Font(size=14, bold=True)

        if gtt_df.empty:
            ws['A3'] = "No confidence data available"
            return

        # Headers
        headers = [
            'Symbol', 'CompositeScore', 'Decision_Confidence', 'Confidence_Level',
            'empirical_bayes_prob', 'adjusted_probability', 'technical_contribution',
            'Confidence_Explanation'
        ]

        for j, header in enumerate(headers, 1):
            cell = ws[f'{get_column_letter(j)}2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Data
        for i, (_, row) in enumerate(gtt_df.iterrows(), 3):
            ws[f'A{i}'] = row.get('Symbol', '')
            ws[f'B{i}'] = row.get('CompositeScore', 0)
            ws[f'C{i}'] = row.get('Decision_Confidence', 0)
            ws[f'D{i}'] = row.get('Confidence_Level', '')
            ws[f'E{i}'] = row.get('empirical_bayes_prob', 0)
            ws[f'F{i}'] = row.get('adjusted_probability', 0)
            ws[f'G{i}'] = row.get('technical_contribution', 0)
            ws[f'H{i}'] = row.get('Confidence_Explanation', '')

        # Conditional formatting
        for i in range(3, len(gtt_df) + 3):
            confidence_cell = ws[f'C{i}']
            if confidence_cell.value >= 4:
                confidence_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            elif confidence_cell.value >= 3:
                confidence_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            else:
                confidence_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        # Auto-adjust column widths
        for j in range(1, len(headers) + 1):
            col_letter = get_column_letter(j)
            ws.column_dimensions[col_letter].width = 18

    def generate_final_excel(self, gtt_df: pd.DataFrame, backtest_results: Dict,
                           output_path: str) -> bool:
        """
        Generate complete final Excel workbook.

        Args:
            gtt_df: GTT plan DataFrame
            backtest_results: Backtest results dictionary
            output_path: Output Excel path

        Returns:
            Success status
        """
        try:
            logger.info("Generating final delivery Excel")

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            self.create_summary_sheet(wb, gtt_df, backtest_results)
            self.create_gtt_plan_sheet(wb, gtt_df)
            self.create_confidence_analysis_sheet(wb, gtt_df)

            # Save workbook
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)

            logger.info(f"Final Excel saved: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to generate final Excel: {str(e)}")
            return False

def run_final_excel(gtt_plan_path: str, output_path: str = "outputs/GTT_Delivery_Final.xlsx",
                   backtest_dir: str = "outputs/backtest") -> bool:
    """
    Run final Excel generation.

    Args:
        gtt_plan_path: Path to GTT plan CSV
        output_path: Output Excel path
        backtest_dir: Backtest results directory

    Returns:
        Success status
    """
    try:
        logger.info("Starting final Excel generation")

        # Initialize generator
        excel_gen = FinalExcelGenerator()

        # Load data
        gtt_df = excel_gen.load_gtt_plan(gtt_plan_path)
        if gtt_df.empty:
            logger.error("No GTT plan data loaded")
            return False

        backtest_results = excel_gen.load_backtest_results(backtest_dir)

        # Generate Excel
        success = excel_gen.generate_final_excel(gtt_df, backtest_results, output_path)

        if success:
            logger.info("Final Excel generation completed successfully")

        return success

    except Exception as e:
        logger.error(f"Final Excel generation failed: {str(e)}")
        return False

# Legacy function for backward compatibility
def build_final_excel(plan_csv: str, out_xlsx: str, tz: str = "Asia/Kolkata") -> None:
    """Legacy function - redirects to new implementation."""
    logger.warning("Using legacy build_final_excel function - consider using run_final_excel instead")
    run_final_excel(plan_csv, out_xlsx)