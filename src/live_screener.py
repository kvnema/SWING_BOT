import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta
import time
import matplotlib.pyplot as plt
import seaborn as sns
import os

# Import instrument lists from config
from .config import INSTRUMENT_KEYS, NIFTY_50_STOCKS, NSE_ETFS, ALL_INSTRUMENTS

def run_live_screener(include_etfs=True):
    instruments_to_screen = ALL_INSTRUMENTS if include_etfs else NIFTY_50_STOCKS
    print(f"Starting Stock Screener for {len(instruments_to_screen)} instruments (ETFs: {include_etfs})...")

    # Upstox API credentials
    API_KEY = os.getenv('UPSTOX_API_KEY', 'b10becd3-d69f-4d44-8ab8-470c0f54c390')
    API_SECRET = os.getenv('UPSTOX_API_SECRET', 'peqijfmcla')
    ACCESS_TOKEN = os.getenv('UPSTOX_ACCESS_TOKEN', 'eyJ0eXAiOiJKV1QiLCJrZXlfaWQiOiJza192MS4wIiwiYWxnIjoiSFMyNTYifQ.eyJzdWIiOiI0V0FRUUIiLCJqdGkiOiI2OTQyOWI1MTgzMWQ0ZjA5NzIxZWJhMzgiLCJpc011bHRpQ2xpZW50IjpmYWxzZSwiaXNQbHVzUGxhbiI6dHJ1ZSwiaWF0IjoxNzY1OTcyODE3LCJpc3MiOiJ1ZGFwaS1nYXRld2F5LXNlcnZpY2UiLCJleHAiOjE3NjYwMDg4MDB9.xAyY8gUzF9BvmmFxCvVYbXEOUImPJNEVSlIeg75ANvA')

    # Use centralized instrument keys
    instrument_keys = INSTRUMENT_KEYS

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    print(f"Fetching live data for {len(instruments_to_screen)} instruments...")

    # Fetch data for all instruments
    all_data = []
    failed_stocks = []

    headers = {
        'Authorization': f'Bearer {ACCESS_TOKEN}',
        'Accept': 'application/json'
    }

    for i, stock in enumerate(instruments_to_screen):
        try:
            print(f"  [{i+1}/{len(instruments_to_screen)}] Fetching {stock}...", end=" ", flush=True)
            
            # Fetch 2 years of historical data
            end_date = datetime.now()
            start_date = end_date - timedelta(days=730)
            
            # Upstox API call for historical data
            symbol = stock.replace('.NS', '')
            instrument_key = instrument_keys.get(symbol, symbol)
            url = f"https://api.upstox.com/v2/historical-candle/{instrument_key}/day/{end_date.strftime('%Y-%m-%d')}/{start_date.strftime('%Y-%m-%d')}"
            
            response = requests.get(url, headers=headers)
            
            if response.status_code == 200:
                data_json = response.json()
                candles = data_json.get('data', {}).get('candles', [])
                
                if candles:
                    df = pd.DataFrame(candles, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'oi'])
                    df['timestamp'] = pd.to_datetime(df['timestamp'])
                    df['Date'] = df['timestamp'].dt.date
                    df['Stock'] = symbol
                    df = df[['Stock', 'Date', 'open', 'high', 'low', 'close', 'volume']]
                    df.columns = ['Stock', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']
                    all_data.append(df)
                    print(f"  [{i+1}/{len(NIFTY_50_STOCKS)}] Fetching {stock}... OK")
                else:
                    print(f"  [{i+1}/{len(NIFTY_50_STOCKS)}] Fetching {stock}... FAILED (No data)")
                    failed_stocks.append(stock)
            else:
                print(f"  [{i+1}/{len(NIFTY_50_STOCKS)}] Fetching {stock}... FAILED (API Error: {response.status_code})")
                failed_stocks.append(stock)
                
        except Exception as e:
            print(f"  [{i+1}/{len(NIFTY_50_STOCKS)}] Fetching {stock}... FAILED (Error: {str(e)[:30]})")
            failed_stocks.append(stock)
        
        time.sleep(0.2)  # Rate limiting

    print(f"\nSuccessfully fetched {len(all_data)} stocks out of {len(NIFTY_50_STOCKS)}")

    if len(all_data) > 0:
        # Combine all data
        df_all = pd.concat(all_data, ignore_index=True)
        df_all = df_all.reset_index()
        df_all = df_all[['Stock', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume']]
        df_all['Date'] = pd.to_datetime(df_all['Date']).dt.date
        df_all = df_all.sort_values(['Stock', 'Date']).reset_index(drop=True)
        
        # Create single comprehensive sheet with all data and indicators
        print("Creating comprehensive data sheet...")
        ws_data = wb.create_sheet('Data_Indicators', 0)
        ws_data.append([
            'Stock', 'Date', 'Open', 'High', 'Low', 'Close', 'Volume',
            'EMA20', 'EMA50', 'EMA200',
            'RS_vs_NIFTY', 'RS_ROC20',
            'RSI14',
            'MACD', 'Signal', 'Histogram',
            'ATR14',
            'Bollinger_Upper', 'Bollinger_Lower', 'Bollinger_Bandwidth',
            'Donchian_20_High', 'Donchian_20_Low',
            'RVOL20',
            'TrendContinuation_Flag', 'Breakout_Flag', 'VCP_Flag', 'SEPA_Flag', 'MeanReversion_Flag',
            'Pivot_R1', 'Pivot_S1', 'Supertrend'
        ])
        
        # Get NIFTY data for RS calculations
        nifty_data = df_all[df_all['Stock'] == 'NIFTY.NS'].set_index('Date')['Close'] if 'NIFTY.NS' in df_all['Stock'].unique() else None
        
        processed_frames = []
        for stock in df_all['Stock'].unique():
            if stock == 'NIFTY.NS':
                continue  # Skip NIFTY itself
                
            stock_data = df_all[df_all['Stock'] == stock].copy()
            
            if len(stock_data) >= 200:  # Need more data for longer indicators
                # Exponential Moving Averages
                stock_data['EMA20'] = stock_data['Close'].ewm(span=20, adjust=False).mean()
                stock_data['EMA50'] = stock_data['Close'].ewm(span=50, adjust=False).mean()
                stock_data['EMA200'] = stock_data['Close'].ewm(span=200, adjust=False).mean()
                
                # RSI14
                delta = stock_data['Close'].diff()
                gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                rs = gain / loss
                stock_data['RSI14'] = 100 - (100 / (1 + rs))
                
                # MACD
                ema12 = stock_data['Close'].ewm(span=12, adjust=False).mean()
                ema26 = stock_data['Close'].ewm(span=26, adjust=False).mean()
                stock_data['MACD'] = ema12 - ema26
                stock_data['Signal'] = stock_data['MACD'].ewm(span=9, adjust=False).mean()
                stock_data['Histogram'] = stock_data['MACD'] - stock_data['Signal']
                
                # ATR14
                high_low = stock_data['High'] - stock_data['Low']
                high_close = (stock_data['High'] - stock_data['Close'].shift(1)).abs()
                low_close = (stock_data['Low'] - stock_data['Close'].shift(1)).abs()
                true_range = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
                stock_data['ATR14'] = true_range.rolling(14).mean()
                
                # Bollinger Bands
                sma20 = stock_data['Close'].rolling(20).mean()
                std = stock_data['Close'].rolling(20).std()
                stock_data['Bollinger_Upper'] = sma20 + (std * 2)
                stock_data['Bollinger_Lower'] = sma20 - (std * 2)
                stock_data['Bollinger_Bandwidth'] = (stock_data['Bollinger_Upper'] - stock_data['Bollinger_Lower']) / sma20
                
                # Donchian Channel
                stock_data['Donchian_20_High'] = stock_data['High'].rolling(20).max()
                stock_data['Donchian_20_Low'] = stock_data['Low'].rolling(20).min()
                
                # RVOL20
                avg_volume_20 = stock_data['Volume'].rolling(20).mean()
                stock_data['RVOL20'] = stock_data['Volume'] / avg_volume_20
                
                # RS vs NIFTY
                if nifty_data is not None:
                    # Align dates
                    common_dates = stock_data.set_index('Date').index.intersection(nifty_data.index)
                    stock_aligned = stock_data.set_index('Date').loc[common_dates]['Close']
                    nifty_aligned = nifty_data.loc[common_dates]
                    stock_data = stock_data.set_index('Date')
                    stock_data['RS_vs_NIFTY'] = stock_aligned / nifty_aligned
                    stock_data['RS_ROC20'] = stock_data['RS_vs_NIFTY'].pct_change(20) * 100
                    stock_data = stock_data.reset_index()
                else:
                    stock_data['RS_vs_NIFTY'] = np.nan
                    stock_data['RS_ROC20'] = np.nan
                
                # Strategy Flags
                # 1. Trend Continuation / Pullback: EMA20 > EMA50 > EMA200, RSI 45-65, price near EMA20
                stock_data['TrendContinuation_Flag'] = (
                    (stock_data['EMA20'] > stock_data['EMA50']) & 
                    (stock_data['EMA50'] > stock_data['EMA200']) & 
                    (stock_data['RSI14'] >= 45) & (stock_data['RSI14'] <= 65) &
                    (abs(stock_data['Close'] - stock_data['EMA20']) < stock_data['ATR14'])
                ).astype(int)
                
                # 2. Breakout: Price > DonchianH20 or BB_Upper, RVOL >=1.5
                stock_data['Breakout_Flag'] = (
                    ((stock_data['Close'] > stock_data['Donchian_20_High'].shift(1)) | 
                     (stock_data['Close'] > stock_data['Bollinger_Upper'].shift(1))) & 
                    (stock_data['RVOL20'] >= 1.5)
                ).astype(int)
                
                # 3. VCP: Falling BB_BandWidth, higher lows, breakout above pivot (using EMA200 as pivot), RVOL spike
                bandwidth_slope = stock_data['Bollinger_Bandwidth'].diff()
                stock_data['VCP_Flag'] = (
                    (bandwidth_slope < 0) & 
                    (stock_data['Low'] > stock_data['Low'].shift(1)) & 
                    (stock_data['Close'] > stock_data['EMA200']) &
                    (stock_data['RVOL20'] > 1.5)
                ).astype(int)
                
                # 4. SEPA: EMAs stacked, low BandWidth, breakout above pivot
                stock_data['SEPA_Flag'] = (
                    (stock_data['Close'] > stock_data['EMA20']) &
                    (stock_data['EMA20'] > stock_data['EMA50']) & 
                    (stock_data['EMA50'] > stock_data['EMA200']) & 
                    (stock_data['Bollinger_Bandwidth'] < 0.1) & 
                    (stock_data['Close'] > stock_data['EMA200'])
                ).astype(int)
                
                # 5. Mean Reversion: In uptrend, Close < EMA20, RSI <35
                stock_data['MeanReversion_Flag'] = (
                    (stock_data['EMA20'] > stock_data['EMA50']) & 
                    (stock_data['EMA50'] > stock_data['EMA200']) &
                    (stock_data['Close'] < stock_data['EMA20']) &
                    (stock_data['RSI14'] < 35)
                ).astype(int)
                
                # 6. Relative Strength Leaders: Will be handled separately for top 10
                
                # Pivot Points (daily)
                stock_data['Pivot'] = (stock_data['High'] + stock_data['Low'] + stock_data['Close']) / 3
                stock_data['Pivot_R1'] = 2 * stock_data['Pivot'] - stock_data['Low']
                stock_data['Pivot_S1'] = 2 * stock_data['Pivot'] - stock_data['High']
                
                # Supertrend (simplified)
                factor = 3
                stock_data['HL2'] = (stock_data['High'] + stock_data['Low']) / 2
                stock_data['ATR_Super'] = stock_data['ATR14']
                stock_data['Upper_Band'] = stock_data['HL2'] + (factor * stock_data['ATR_Super'])
                stock_data['Lower_Band'] = stock_data['HL2'] - (factor * stock_data['ATR_Super'])
                stock_data['Supertrend'] = stock_data['Upper_Band']
                for i in range(1, len(stock_data)):
                    if stock_data.loc[stock_data.index[i], 'Close'] > stock_data.loc[stock_data.index[i-1], 'Supertrend']:
                        stock_data.loc[stock_data.index[i], 'Supertrend'] = max(stock_data.loc[stock_data.index[i], 'Lower_Band'], stock_data.loc[stock_data.index[i-1], 'Supertrend'])
                    else:
                        stock_data.loc[stock_data.index[i], 'Supertrend'] = min(stock_data.loc[stock_data.index[i], 'Upper_Band'], stock_data.loc[stock_data.index[i-1], 'Supertrend'])
                
                for idx, row in stock_data.iterrows():
                    ws_data.append([
                        row['Stock'],
                        str(row['Date']),
                        round(row['Open'], 2) if pd.notna(row['Open']) else '',
                        round(row['High'], 2) if pd.notna(row['High']) else '',
                        round(row['Low'], 2) if pd.notna(row['Low']) else '',
                        round(row['Close'], 2) if pd.notna(row['Close']) else '',
                        int(row['Volume']) if pd.notna(row['Volume']) else '',
                        round(row['EMA20'], 2) if pd.notna(row['EMA20']) else '',
                        round(row['EMA50'], 2) if pd.notna(row['EMA50']) else '',
                        round(row['EMA200'], 2) if pd.notna(row['EMA200']) else '',
                        round(row['RS_vs_NIFTY'], 4) if pd.notna(row['RS_vs_NIFTY']) else '',
                        round(row['RS_ROC20'], 2) if pd.notna(row['RS_ROC20']) else '',
                        round(row['RSI14'], 2) if pd.notna(row['RSI14']) else '',
                        round(row['MACD'], 4) if pd.notna(row['MACD']) else '',
                        round(row['Signal'], 4) if pd.notna(row['Signal']) else '',
                        round(row['Histogram'], 4) if pd.notna(row['Histogram']) else '',
                        round(row['ATR14'], 2) if pd.notna(row['ATR14']) else '',
                        round(row['Bollinger_Upper'], 2) if pd.notna(row['Bollinger_Upper']) else '',
                        round(row['Bollinger_Lower'], 2) if pd.notna(row['Bollinger_Lower']) else '',
                        round(row['Bollinger_Bandwidth'], 4) if pd.notna(row['Bollinger_Bandwidth']) else '',
                        round(row['Donchian_20_High'], 2) if pd.notna(row['Donchian_20_High']) else '',
                        round(row['Donchian_20_Low'], 2) if pd.notna(row['Donchian_20_Low']) else '',
                        round(row['RVOL20'], 2) if pd.notna(row['RVOL20']) else '',
                        int(row['TrendContinuation_Flag']) if pd.notna(row['TrendContinuation_Flag']) else '',
                        int(row['Breakout_Flag']) if pd.notna(row['Breakout_Flag']) else '',
                        int(row['VCP_Flag']) if pd.notna(row['VCP_Flag']) else '',
                        int(row['SEPA_Flag']) if pd.notna(row['SEPA_Flag']) else '',
                        int(row['MeanReversion_Flag']) if pd.notna(row['MeanReversion_Flag']) else '',
                        round(row['Pivot_R1'], 2) if pd.notna(row['Pivot_R1']) else '',
                        round(row['Pivot_S1'], 2) if pd.notna(row['Pivot_S1']) else '',
                        round(row['Supertrend'], 2) if pd.notna(row['Supertrend']) else ''
                    ])
                # keep processed stock dataframe for later screener/backtest outputs
                processed_frames.append(stock_data)
        
        ws_data.freeze_panes = 'A2'
        # Set column widths for all columns (31 columns: A to AE)
        for col_num in range(1, 32):  # 1 to 31
            col_letter = get_column_letter(col_num)
            ws_data.column_dimensions[col_letter].width = 12
        
        # Save the workbook with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'NIFTY50_Screener_Live_{timestamp}.xlsx'
        print(f"Saving Excel file as {filename}...")
        wb.save(filename)
        print(f"Excel file saved successfully: {filename}")
        # Build a consolidated processed dataframe containing computed indicators & flags
        if processed_frames:
            df_processed = pd.concat(processed_frames, ignore_index=True)
        else:
            df_processed = pd.DataFrame()

        print(f"Total records processed: {len(df_processed)}")
        print(f"Stocks with sufficient data: {len(df_processed['Stock'].unique()) if not df_processed.empty else 0}")
        
        # Save full processed data for new project
        if not df_processed.empty:
            df_processed.to_csv('nifty50_indicators_full.csv', index=False)
            print("Full indicators data saved: nifty50_indicators_full.csv")
        
        # Create screener for latest date
        # Use processed dataframe (contains flags and indicators) for screener outputs
        if df_processed.empty:
            print('No processed data available to build screener.')
            wb.save(filename)
            return

        latest_date = df_processed['Date'].max()
        latest_data = df_processed[df_processed['Date'] == latest_date].copy()
        
        # Add CompositeScore: sum of flags
        latest_data['CompositeScore'] = latest_data[['TrendContinuation_Flag', 'Breakout_Flag', 'VCP_Flag', 'SEPA_Flag', 'MeanReversion_Flag']].sum(axis=1)
        
        # For RS Leaders, add a flag for top 10
        top_rs = latest_data.nlargest(10, 'RS_ROC20')['Stock']
        latest_data['RS_Leader_Flag'] = latest_data['Stock'].isin(top_rs).astype(int)
        
        # Screener columns: Symbol | Close | EMA20 | EMA50 | EMA200 | RSI14 | ATR14 | RVOL20 | RS_ROC20 | CompositeScore | TrendContinuation | VCP | SEPA | Breakout | MeanReversion
        screener_cols = ['Stock', 'Close', 'EMA20', 'EMA50', 'EMA200', 'RSI14', 'ATR14', 'RVOL20', 'RS_ROC20', 'CompositeScore', 'TrendContinuation_Flag', 'VCP_Flag', 'SEPA_Flag', 'Breakout_Flag', 'MeanReversion_Flag']
        screener_df = latest_data[screener_cols]
        screener_df.to_csv(f'screener_latest_{timestamp}.csv', index=False)
        print(f"Screener CSV saved: screener_latest_{timestamp}.csv")
        
        # Ranked Top-25
        ranked_df = latest_data.sort_values('CompositeScore', ascending=False).head(25)
        ranked_df.to_csv(f'ranked_top25_{timestamp}.csv', index=False)
        print(f"Ranked Top-25 CSV saved: ranked_top25_{timestamp}.csv")
        
        # Visualizations
        # Bar plot of RS_ROC20
        plt.figure(figsize=(12, 8))
        latest_data_sorted = latest_data.sort_values('RS_ROC20', ascending=False)
        plt.bar(latest_data_sorted['Stock'], latest_data_sorted['RS_ROC20'])
        plt.title('RS_ROC20 by Stock')
        plt.xticks(rotation=90)
        plt.tight_layout()
        plt.savefig(f'rs_roc_bar_{timestamp}.png')
        print(f"RS_ROC20 bar plot saved: rs_roc_bar_{timestamp}.png")
        
        # Trend charts for top 5 candidates (use processed data)
        top_stocks = ranked_df['Stock'].head(5)
        for stock in top_stocks:
            stock_hist = df_processed[df_processed['Stock'] == stock].tail(100)  # last 100 days
            plt.figure(figsize=(12, 6))
            plt.plot(stock_hist['Date'], stock_hist['Close'], label='Close')
            plt.plot(stock_hist['Date'], stock_hist['EMA20'], label='EMA20')
            plt.plot(stock_hist['Date'], stock_hist['EMA50'], label='EMA50')
            plt.plot(stock_hist['Date'], stock_hist['EMA200'], label='EMA200')
            plt.title(f'{stock} Trend Chart (Last 100 Days)')
            plt.legend()
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(f'{stock}_trend_{timestamp}.png')
            print(f"Trend chart for {stock} saved: {stock}_trend_{timestamp}.png")
        
        # Backtest
        strategies = {
            'TrendContinuation': 'TrendContinuation_Flag',
            'Breakout': 'Breakout_Flag',
            'VCP': 'VCP_Flag',
            'SEPA': 'SEPA_Flag',
            'MeanReversion': 'MeanReversion_Flag'
        }
        backtest_results = {}
        for strat_name, flag_col in strategies.items():
            returns = []
            for stock in df_all['Stock'].unique():
                if stock == 'NIFTY.NS':
                    continue
                stock_df = df_processed[df_processed['Stock'] == stock].set_index('Date')
                entry_dates = stock_df[stock_df[flag_col] == 1].index
                for entry_date in entry_dates:
                    exit_date = entry_date + timedelta(days=20)
                    if exit_date in stock_df.index:
                        entry_price = stock_df.loc[entry_date, 'Close']
                        exit_price = stock_df.loc[exit_date, 'Close']
                        ret = (exit_price - entry_price) / entry_price * 100
                        returns.append(ret)
            if returns:
                win_rate = sum(1 for r in returns if r > 0) / len(returns) * 100
                avg_return = sum(returns) / len(returns)
                backtest_results[strat_name] = {'Win_Rate_%': win_rate, 'Avg_Return_%': avg_return, 'Total_Trades': len(returns)}
            else:
                backtest_results[strat_name] = {'Win_Rate_%': 0, 'Avg_Return_%': 0, 'Total_Trades': 0}
        
        backtest_df = pd.DataFrame(backtest_results).T
        backtest_df.to_csv(f'backtest_results_{timestamp}.csv')
        print(f"Backtest results saved: backtest_results_{timestamp}.csv")
        print("Backtest Summary:")
        print(backtest_df)

    else:
        print("✗ Failed to fetch data for any stocks. Please check your internet connection.")

if __name__ == '__main__':
    run_live_screener()

