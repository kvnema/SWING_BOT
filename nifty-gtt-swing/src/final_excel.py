import pandas as pd
from datetime import datetime
import pytz


def build_final_excel(plan_csv: str, out_xlsx: str, tz: str = "Asia/Kolkata") -> None:
    """Build final Excel from GTT plan CSV for Delivery mode."""
    try:
        df = pd.read_csv(plan_csv)
    except (pd.errors.EmptyDataError, FileNotFoundError):
        df = pd.DataFrame()  # Empty plan

    if df.empty:
        # Create empty Excel with banner
        final_df = pd.DataFrame(columns=['Symbol', 'GTT_Buy_Price', 'Stoploss', 'Sell_Rate', 'Strategy', 'Notes', 'Explanation', 'DecisionConfidence', 'CI_low', 'CI_high', 'OOS_WinRate', 'OOS_ExpectancyR', 'Trades_OOS', 'Confidence_Reason', 'RSI14', 'RSI14_Status', 'GoldenBull_Flag', 'GoldenBear_Flag', 'GoldenBull_Date', 'GoldenBear_Date', 'Generated_At_IST'])
        tz_obj = pytz.timezone(tz)
        now_ist = datetime.now(tz_obj).strftime('%Y-%m-%d %H:%M:%S')
        final_df.loc[0] = ['No actionable signals today', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', now_ist]
    else:
        # Transform to required columns
        final_df = pd.DataFrame({
            'Symbol': df['Symbol'],
            'GTT_Buy_Price': df['ENTRY_trigger_price'],
            'Stoploss': df['STOPLOSS_trigger_price'],
            'Sell_Rate': df['TARGET_trigger_price'],
            'Strategy': df['Strategy'],
            'Notes': df['Notes'],
            'Explanation': df['Explanation'],
            'DecisionConfidence': df.get('DecisionConfidence', 0.5),
            'CI_low': df.get('CI_low', 0.4),
            'CI_high': df.get('CI_high', 0.6),
            'OOS_WinRate': df.get('OOS_WinRate', 0.5),
            'OOS_ExpectancyR': df.get('OOS_ExpectancyR', 0.0),
            'Trades_OOS': df.get('Trades_OOS', 0),
            'Confidence_Reason': df.get('Confidence_Reason', ''),
            'RSI14': df['RSI14'],
            'RSI14_Status': df['RSI14_Status'],
            'GoldenBull_Flag': df['GoldenBull_Flag'],
            'GoldenBear_Flag': df['GoldenBear_Flag'],
            'GoldenBull_Date': df['GoldenBull_Date'],
            'GoldenBear_Date': df['GoldenBear_Date']
        })
        
        # Add timestamp
        tz_obj = pytz.timezone(tz)
        now_ist = datetime.now(tz_obj).strftime('%Y-%m-%d %H:%M:%S')
        final_df['Generated_At_IST'] = now_ist
        
        # Sort by Strategy then Symbol
        final_df = final_df.sort_values(['Strategy', 'Symbol']).reset_index(drop=True)
    
    # Write to Excel
    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name='GTT-Delivery-Plan', index=False)
        
        # Format percentage columns
        workbook = writer.book
        worksheet = writer.sheets['GTT-Delivery-Plan']
        
        # Find column indices for percentage formatting
        col_names = list(final_df.columns)
        decision_conf_col = col_names.index('DecisionConfidence') + 1  # 1-based
        ci_low_col = col_names.index('CI_low') + 1
        ci_high_col = col_names.index('CI_high') + 1
        oos_winrate_col = col_names.index('OOS_WinRate') + 1
        
        # Apply percentage format to data rows (skip header)
        from openpyxl.styles import NamedStyle
        pct_style = NamedStyle(name='percentage', number_format='0.0%')
        if 'percentage' not in workbook.named_styles:
            workbook.add_named_style(pct_style)
        
        for row in range(2, len(final_df) + 2):  # Start from row 2 (after header)
            worksheet.cell(row=row, column=decision_conf_col).style = 'percentage'
            worksheet.cell(row=row, column=ci_low_col).style = 'percentage'
            worksheet.cell(row=row, column=ci_high_col).style = 'percentage'
            worksheet.cell(row=row, column=oos_winrate_col).style = 'percentage'
        
        # Add footer
        last_row = len(final_df) + 2  # +1 for header, +1 for blank
        
        # Footer
        worksheet.cell(row=last_row, column=1).value = "Product: Delivery (D)"
        worksheet.cell(row=last_row + 1, column=1).value = "ENTRY trigger type per strategy: MR → BELOW EMA20; Breakout → ABOVE DonchianH20"
        worksheet.cell(row=last_row + 2, column=1).value = "Ensure EDIS authorization is active for SELL legs"